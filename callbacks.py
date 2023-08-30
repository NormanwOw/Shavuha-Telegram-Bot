from openpyxl import Workbook
from aiogram.types import LabeledPrice
import aiogram.utils.exceptions
from openpyxl.styles import Alignment

from config import PAY_TOKEN
from messages import *
from functions import *
from markups import *
import pages


async def basket(user_id: int, callback: types.CallbackQuery, bot: Bot):
    if await OrderDB.get_order_by_id(user_id) is None:
        await callback.answer('Корзина пуста')
    else:
        await bot.send_message(user_id, await basket_title(user_id),
                               reply_markup=await pages.basket_menu_page(user_id))


async def add_to_basket(user_id: int, msg_id: int, callback: types.CallbackQuery, bot: Bot):
    product = callback.data[11:]
    await OrderDB.temp_to_order(user_id)
    await OrderDB.set_order_time(user_id, get_time()[0])
    await callback.answer('Товар добавлен в корзину')
    await bot.edit_message_reply_markup(user_id, msg_id, reply_markup=await pages.product_page(user_id, product))


async def back_to_basket(user_id: int, msg_id: int, callback: types.CallbackQuery, bot: Bot):
    if 'del' in callback.data:
        await OrderDB.delete_comment(user_id)
        await callback.answer('Комментарий удалён')
    await bot.edit_message_text(await basket_title(user_id), user_id, msg_id,
                                reply_markup=await pages.basket_menu_page(user_id))


async def basket_product_counter(user_id: int, msg_id: int, callback: types.CallbackQuery, bot: Bot):
    product = callback.data[4:]
    if '!up' in callback.data:
        count = 1
    else:
        count = -1
    await OrderDB.add_order(user_id, {product: count})

    if await OrderDB.get_products_count(user_id) == 0:
        await OrderDB.clear_basket(user_id)
        await bot.edit_message_text(EMPTY_BASKET, user_id, msg_id,
                                    reply_markup=await pages.basket_menu_page(user_id))
        return

    await bot.edit_message_reply_markup(user_id, msg_id, reply_markup=await pages.basket_menu_page(user_id))


async def order_product_counter(user_id: int, msg_id: int, callback: types.CallbackQuery, bot: Bot):
    count = await OrderDB.get_count(user_id)
    product = callback.data[1:]
    if await OrderDB.is_temp_exists(user_id) is False:
        await OrderDB.add_temp(user_id, product)
    else:
        if '+' in callback.data:
            if count > 21:
                await callback.answer()
                return
            await OrderDB.set_count(user_id, count + 1)
        else:
            if count == 1:
                await callback.answer()
                return
            await OrderDB.set_count(user_id, count - 1)

    await bot.edit_message_reply_markup(user_id, msg_id, reply_markup=await pages.product_page(user_id, product))


async def set_order_comment(user_id: int, msg_id: int, bot: Bot, comment):
    if await OrderDB.get_comment(user_id) is None:
        await bot.send_message(user_id, 'Комментарий к заказу:', reply_markup=ikb_cancel)
        await comment.get_comment.set()
    else:
        await bot.edit_message_text(await order_comment_title(user_id), user_id, msg_id,
                                    reply_markup=await pages.comment_page())


async def set_order_time(user_id: int, msg_id: int, callback: types.CallbackQuery, bot: Bot):
    if callback.data == 'set_time':
        _, hour, minute = get_time()
        await bot.edit_message_text(SET_TIME_MESSAGE, user_id, msg_id,
                                    reply_markup=await pages.set_time_page(user_id, hour, minute))
        await OrderDB.set_order_time(user_id, get_time()[0])
    elif callback.data == 'cancel_set_time':
        if await OrderDB.get_order_user_time(user_id) is None:
            await callback.answer()
            return

        await OrderDB.set_order_user_time(user_id, None)
        await bot.edit_message_text(await basket_title(user_id), user_id, msg_id,
                                    reply_markup=await pages.basket_menu_page(user_id))

        await callback.answer('Точное время отменено')
    else:
        time = callback.data[9:]
        await callback.answer(time)
        await OrderDB.set_order_user_time(user_id, time)
        await bot.edit_message_text(await basket_title(user_id), user_id, msg_id,
                                    reply_markup=await pages.basket_menu_page(user_id))


async def set_order_time_navigation(user_id: int, msg_id: int, callback: types.CallbackQuery, bot: Bot):
    _, hour, minute = get_time()
    if 'next' in callback.data:
        hour += 6
    try:
        await bot.edit_message_reply_markup(user_id, msg_id,
                                            reply_markup=await pages.set_time_page(user_id, hour, minute))
    except aiogram.utils.exceptions.MessageNotModified:
        await callback.answer()
        return


async def create_invoice(user_id: int, msg_id: int, bot: Bot):
    data = await OrderDB.get_order_by_id(user_id)
    try:
        order_list = json.loads(data[2])
    except TypeError:
        await bot.delete_message(user_id, msg_id)
        return
    order_prices = []

    prices = await OrderDB.get_prices()
    desc = ''
    for item in order_list:
        for i in prices:
            if i[0] == item:
                p = i[1]
        cnt = order_list[item]
        label = item + f' - {cnt}шт.'
        lp = LabeledPrice(label=label, amount=p * cnt * 100)
        order_prices.append(lp)

        desc += f' ▫️ {item}'

    await bot.send_invoice(user_id,
                           title='Заказ',
                           description=desc,
                           provider_token=PAY_TOKEN,
                           currency='rub',
                           is_flexible=True,
                           start_parameter='example',
                           payload=order_list,
                           prices=order_prices,
                           need_shipping_address=False)


async def edit_menu_navigation(user_id: int, msg_id: int, callback: types.CallbackQuery, bot: Bot, del_product: bool):
    data = callback.data.split()
    page = int(data[1])
    next_page_len = int(data[2])
    if 'next' in callback.data and next_page_len > 1:
        await bot.edit_message_text(EDIT_MENU_TITLE, user_id, msg_id,
                                    reply_markup=await pages.edit_menu_page(del_product, page + 1))
    elif 'prev' in callback.data and page != 1:
        await bot.edit_message_text(EDIT_MENU_TITLE, user_id, msg_id,
                                    reply_markup=await pages.edit_menu_page(del_product, page - 1))


async def get_xlsx() -> str:
    wb = Workbook()
    ws = wb.active
    ws.append(['Номер заказа', 'Заказ', 'Стоимость', 'Дата', 'Время'])

    table = ['A', 'B', 'C', 'D', 'E']

    for ch in table:
        cell = ws[f'{ch}1']
        cell.style = 'Accent1'
        cell.alignment = Alignment(horizontal='center')

    for i, order in enumerate(await OrderDB.get_all_from_archive()):
        ws.append([order[i] for i in [1, 4, 3, 6, 7]])
        ws[f'C{i + 2}'].number_format = '#,## ₽'
        ws[f'D{i + 2}'].alignment = Alignment(horizontal='right')
        ws[f'E{i + 2}'].alignment = Alignment(horizontal='right')

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 10

    now = datetime.datetime.now().strftime('%d.%m.%Y')
    name = now+'.xlsx'
    wb.save(name)

    return name
