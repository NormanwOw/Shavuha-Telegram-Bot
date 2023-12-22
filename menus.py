import os
from abc import ABC, abstractmethod

from aiogram.types import CallbackQuery, LabeledPrice
import aiogram.utils.exceptions
from aiogram import Bot
from openpyxl import Workbook
from openpyxl.styles import Alignment

from config import PAY_TOKEN
from functions import *
from messages import *
from markups import *
from states import *


class Menu(ABC):
    db = OrderDB

    @classmethod
    @abstractmethod
    async def get_page(cls, *args):
        pass


class Basket(Menu):

    @classmethod
    async def get_page(cls, user_id: int) -> InlineKeyboardMarkup:
        prices = await cls.db.get_prices()
        order = await cls.db.get_order_list(user_id)
        ikb = InlineKeyboardMarkup(row_width=3)
        total_price = 0
        try:
            for item in order:
                for price in prices:
                    if item == price[0]:
                        p = order[item] * price[1]
                        total_price += p
                ikb.row(InlineKeyboardButton(f'➖       {item}',
                                             callback_data=f'!dn_{item}'),
                        InlineKeyboardButton(f'[{order[item]}шт.]          ➕',
                                             callback_data=f'!up_{item}'))
        except TypeError:
            return menu

        ikb.add(InlineKeyboardButton('Указать время', callback_data='set_time'))
        ikb.insert(InlineKeyboardButton('Комментарий', callback_data='order_comment'))
        ikb.add(InlineKeyboardButton('Меню', switch_inline_query_current_chat='#menu'))
        ikb.add(InlineKeyboardButton(f'Оплатить {total_price},00 RUB', callback_data='pay'))

        return ikb

    @classmethod
    async def show_page(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot,
                        check_basket: bool = False):
        if check_basket:
            if await cls.db.get_order_by_id(user_id) is None:
                await callback.answer('Корзина пуста')
            else:
                await bot.send_message(
                    chat_id=user_id,
                    text=await cls.get_title(user_id),
                    reply_markup=await cls.get_page(user_id)
                )
        else:
            if 'del' in callback.data:
                await cls.db.delete_comment(user_id)
                await callback.answer('Комментарий удалён')

            await bot.edit_message_text(
                text=await cls.get_title(user_id),
                chat_id=user_id,
                message_id=msg_id,
                reply_markup=await cls.get_page(user_id)
            )

    @classmethod
    async def set_time_page(cls, user_id: int, hour: int, minute: int) -> InlineKeyboardMarkup:
        ikb = InlineKeyboardMarkup(row_width=6)
        if hour > 23:
            hour -= 24

        if 0 <= minute < 15:
            minute = 15
        elif 15 <= minute < 30:
            minute = 30
        elif 30 <= minute < 45:
            minute = 45
        elif 45 <= minute < 60:
            minute = 0
            hour += 1

        hour_var = 0
        minute_var = 0

        for i in range(24):
            if hour + hour_var < 10:
                h = f'0{hour + hour_var}'
            else:
                h = hour + hour_var
            if minute + minute_var == 0:
                m = '00'
            else:
                m = minute + minute_var

            if hour + hour_var == 24:
                h = '00'

            ikb.insert(InlineKeyboardButton(
                text=f'{h}:{m}',
                callback_data=f'set_time_{h}:{m}'
                )
            )
            if hour + hour_var > 23:
                hour = 0
                hour_var = 0

            minute_var += 15
            if minute + minute_var > 45:
                minute = 0
                minute_var = 0
                hour_var += 1

        ikb.add(InlineKeyboardButton('◀️', callback_data='prev_time_page'))
        ikb.insert(InlineKeyboardButton('▶️', callback_data='next_time_page'))
        order_user_time = await cls.db.get_order_user_time(user_id)
        if order_user_time is not None:
            ikb.add(InlineKeyboardButton(f'[{order_user_time}] Отменить',
                                         callback_data='cancel_set_time'))
        ikb.add(InlineKeyboardButton('Назад',
                                     callback_data='back_to_basket'))

        return ikb

    @classmethod
    async def show_time_page(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        _, hour, minute = await get_time()
        if 'next' in callback.data:
            hour += 6
        try:
            await bot.edit_message_reply_markup(
                chat_id=user_id,
                message_id=msg_id,
                reply_markup=await cls.set_time_page(user_id, hour, minute)
            )

        except aiogram.utils.exceptions.MessageNotModified:
            await callback.answer()
            return

    @classmethod
    async def get_title(cls, user_id: int) -> str:
        title = '🛒 <b>Корзина</b>'
        time = await cls.db.get_order_user_time(user_id)
        comment = await cls.db.get_comment(user_id)
        if time is None:
            time = ''
        else:
            time = '\n⏱ ' + await cls.db.get_order_user_time(user_id)

        if comment is None:
            comment = ''
        else:
            comment = '\n✏ ' + comment

        return title + time + comment

    @classmethod
    async def set_time(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        if callback.data == 'set_time':
            time, hour, minute = await get_time()

            await bot.edit_message_text(
                text=SET_TIME_MESSAGE,
                chat_id=user_id,
                message_id=msg_id,
                reply_markup=await cls.set_time_page(user_id, hour, minute)
            )

            await cls.db.set_order_time(user_id, time)

        elif callback.data == 'cancel_set_time':
            if await cls.db.get_order_user_time(user_id) is None:
                await callback.answer()
                return

            await cls.db.set_order_user_time(user_id, None)
            await cls.show_page(user_id, msg_id, callback, bot)

            await callback.answer('Точное время отменено')
        else:
            time = callback.data[9:]
            await callback.answer(time)
            await cls.db.set_order_user_time(user_id, time)
            await cls.show_page(user_id, msg_id, callback, bot)

    @classmethod
    async def get_comment_title(cls, user_id: int) -> str:
        title = '<b>✏ Комментарий к заказу</b>'
        comment = await cls.db.get_comment(user_id)
        if comment is None:
            comment = ''
        else:
            comment = '\n ' + comment
        return title + comment

    @classmethod
    async def set_comment(cls, user_id: int, msg_id: int, bot: Bot):
        if await cls.db.get_comment(user_id) is None:

            await bot.send_message(
                chat_id=user_id,
                text='Комментарий к заказу:',
                reply_markup=ikb_cancel
            )

            await OrderComment.get_comment.set()
        else:
            await bot.edit_message_text(
                text=await cls.get_comment_title(user_id),
                chat_id=user_id,
                message_id=msg_id,
                reply_markup=await cls.comment_page()
            )

    @staticmethod
    async def comment_page() -> InlineKeyboardMarkup:
        ikb = InlineKeyboardMarkup()
        ikb.add(InlineKeyboardButton('Назад', callback_data='back_to_basket'))
        ikb.insert(InlineKeyboardButton('Удалить', callback_data='del_back_to_basket'))

        return ikb

    @classmethod
    async def product_counter(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        product = callback.data[4:]
        if '!up' in callback.data:
            count = 1
        else:
            count = -1
        await cls.db.add_order(user_id, {product: count})

        if await cls.db.get_products_count(user_id) == 0:
            await cls.db.clear_basket(user_id)

            await bot.edit_message_text(
                text=EMPTY_BASKET,
                chat_id=user_id,
                message_id=msg_id,
                reply_markup=await cls.get_page(user_id)
            )

            return

        await bot.edit_message_reply_markup(
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=await Basket.get_page(user_id)
        )

    @classmethod
    async def get_pay_invoice(cls, user_id: int, msg_id: int, bot: Bot):
        data = await cls.db.get_order_by_id(user_id)
        try:
            order_list = json.loads(data[2])
        except TypeError:
            await bot.delete_message(user_id, msg_id)
            return
        order_prices = []

        prices = await cls.db.get_prices()
        desc = ''
        p = ''
        for item in order_list:
            for i in prices:
                if i[0] == item:
                    p = i[1]
            cnt = order_list[item]
            label = item + f' - {cnt}шт.'
            lp = LabeledPrice(label=label, amount=p * cnt * 100)
            order_prices.append(lp)

            desc += f' ▫️ {item}'

        await bot.send_invoice(
            chat_id=user_id,
            title='Заказ',
            description=desc,
            provider_token=PAY_TOKEN,
            currency='rub',
            start_parameter='example',
            payload=order_list,
            prices=order_prices,
            need_shipping_address=False
        )


class Product(Menu):

    @classmethod
    async def get_page(cls, user_id: int, product: str) -> InlineKeyboardMarkup:
        ikb = InlineKeyboardMarkup(row_width=3)
        ikb.row(InlineKeyboardButton('-', callback_data=f'-{product}'),
                InlineKeyboardButton(f'{await cls.db.get_count(user_id)}', callback_data='None'),
                InlineKeyboardButton('+', callback_data=f'+{product}'))
        ikb.add(InlineKeyboardButton('Добавить в корзину', callback_data=f'basket_add_{product}'))
        products_count = await cls.db.get_products_count(user_id)
        if products_count == 0:
            ikb.add(InlineKeyboardButton('Корзина', callback_data='basket'))
        else:
            ikb.add(InlineKeyboardButton(f'Корзина ({products_count})', callback_data='basket'))
        ikb.insert(InlineKeyboardButton('Меню', switch_inline_query_current_chat='#menu'))

        return ikb

    @classmethod
    async def product_counter(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        count = await cls.db.get_count(user_id)
        product = callback.data[1:]
        if await cls.db.is_temp_exists(user_id) is False:
            await cls.db.add_temp(user_id, product)
        else:
            if '+' in callback.data:
                if count > 21:
                    await callback.answer()
                    return
                await cls.db.set_count(user_id, count + 1)
            else:
                if count == 1:
                    await callback.answer()
                    return
                await cls.db.set_count(user_id, count - 1)

        await bot.edit_message_reply_markup(
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=await cls.get_page(user_id, product)
        )

    @classmethod
    async def add(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        product = callback.data[11:]
        time = await get_time()
        await cls.db.temp_to_order(user_id)
        await cls.db.set_order_time(user_id, time[0])
        await callback.answer('Товар добавлен в корзину')

        await bot.edit_message_reply_markup(
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=await cls.get_page(user_id, product)
        )


class Admin(Menu):

    @staticmethod
    async def get_page() -> InlineKeyboardMarkup:
        ikb = InlineKeyboardMarkup()
        ikb.row(InlineKeyboardButton('Персонал', callback_data='admin_employees'),
                InlineKeyboardButton('Редактор меню', callback_data='admin_menu'))
        ikb.row(InlineKeyboardButton('Статистика', callback_data='admin_stats'),
                InlineKeyboardButton('Рассылки', callback_data='admin_mails'))
        ikb.add(InlineKeyboardButton('История заказов .xlsx', callback_data='admin_xlsx'))
        ikb.add(InlineKeyboardButton('Сообщить об ошибке', callback_data='admin_error'))
        ikb.add(InlineKeyboardButton('Настройки', callback_data='admin_settings'))

        return ikb

    @classmethod
    async def show_page(cls, user_id: int, msg_id: int, bot: Bot, show_stats: bool = False):
        if show_stats:
            avg_price = await cls.db.get_avg_order_price()
            orders_day = await cls.db.get_orders_count_day()
            orders_month = await cls.db.get_orders_count_month()
            orders = await cls.db.get_orders_count()
            stats = f'\n ▫️ Заказов за 24 часа: {orders_day}' \
                    f'\n ▫️ Заказов за 30 дней: {orders_month}' \
                    f'\n ▫️ Всего заказов: {orders}' \
                    f'\n ▫️ Средний чек: {int(avg_price)}'
        else:
            stats = ''

        await bot.edit_message_text(
            text=ADMIN_TITLE + stats,
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=await cls.get_page()
        )

    @classmethod
    async def get_xlsx(cls, user_id: int, bot: Bot):
        wb = Workbook()
        ws = wb.active
        ws.append(['Номер заказа', 'Заказ', 'Стоимость', 'Дата', 'Время'])

        table = ['A', 'B', 'C', 'D', 'E']

        for ch in table:
            cell = ws[f'{ch}1']
            cell.style = 'Accent1'
            cell.alignment = Alignment(horizontal='center')

        archive = await cls.db.get_all_from_archive()

        for i, order in enumerate(archive):
            ws.append([order[i] for i in [1, 4, 3, 6, 7]])
            ws[f'C{i + 2}'].number_format = '#,## ₽'
            ws[f'D{i + 2}'].alignment = Alignment(horizontal='right')
            ws[f'E{i + 2}'].alignment = Alignment(horizontal='right')

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 13
        ws.column_dimensions["E"].width = 10

        now = datetime.datetime.now() + datetime.timedelta(hours=TIME_ZONE)
        path = now.strftime('%d.%m.%Y') + '.xlsx'
        wb.save(path)

        await bot.send_document(user_id, open(path, 'rb'))
        for file in os.listdir():
            if '.xlsx' in file:
                os.remove(file)

    @classmethod
    async def get_error(cls, callback: CallbackQuery):
        await callback.message.answer(
            text=ERROR_TITLE,
            reply_markup=ikb_cancel
        )
        await ErrorHandler.get_error.set()


class Employees(Menu):

    @classmethod
    async def get_page(cls) -> InlineKeyboardMarkup:
        ikb = InlineKeyboardMarkup(row_width=3)
        employees_list = await cls.db.get_id_name_by_status('Повар')
        for employee in employees_list:
            ikb.add(InlineKeyboardButton('🚫', callback_data=f'delete_employee_{employee[0]}'))
            ikb.insert(InlineKeyboardButton(f'{employee[1]}', callback_data='None'))

        ikb.add(InlineKeyboardButton('Сменить пароль', callback_data='update_password'))
        ikb.add(InlineKeyboardButton('Назад', callback_data='back'))
        ikb.insert(InlineKeyboardButton('Помощь', callback_data='employee_help'))

        return ikb

    @classmethod
    async def show_page(cls, user_id: int, msg_id: int, bot: Bot,
                        update_pw: bool = False, show_help: bool = False):
        if update_pw:
            pw = await update_password()
        else:
            password = await get_json('data.json')
            pw = password["employee_password"]

        help_msg = EMPLOYEE_HELP if show_help else ''

        await bot.edit_message_text(
            text=EMPLOYEE_TITLE + f'\nПароль для персонала: <b>{pw}</b>{help_msg}',
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=await cls.get_page()
        )

    @classmethod
    async def delete(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        employee_id = int(callback.data[16:])
        await cls.db.delete_employee(employee_id)
        await cls.show_page(user_id, msg_id, bot)

    @classmethod
    async def send_order_to_employees(
            cls, comment: str, order_list: str, bot: Bot, order_number: int, user_time_str: str,
            price: int, date: str, time: str
    ):

        if comment is None:
            comm = ''
        else:
            comm = f'\n\n✏ Комментарий: {comment}'

        order = json.loads(order_list.replace('\'', '"'))
        order_str = ''
        for employee in await cls.db.get_id_by_status('Повар'):
            for product in order:
                order_str += f'\n ▫️ {product}: {order[product]}'

            await bot.send_message(
                chat_id=employee,
                text=f'<b>Заказ №<u>{order_number}</u></b>' + user_time_str + order_str +
                     f'\n__________\n'
                     f'{price}₽' + comm + f'\n\n'
                     f'{date} {time}'
            )


class EditMenu(Menu):

    @classmethod
    async def get_page(cls, del_product: bool, page: int = 1) -> InlineKeyboardMarkup:
        rows = 6
        ikb = InlineKeyboardMarkup(row_width=3)
        prices_list = await cls.db.get_prices()
        prices_list_rows = prices_list[rows * page - rows:rows * page]
        next_page_len = len(prices_list) - rows * page

        if next_page_len < 0:
            next_page_len = 0

        for product, price, desc, url in prices_list_rows:
            if del_product:
                ikb.add(InlineKeyboardButton('🚫', callback_data='delete_product_' + product))
                ikb.insert(InlineKeyboardButton(product, callback_data='None'))
            else:
                ikb.add(InlineKeyboardButton('✏ ', callback_data='change_desc_' + product))
                ikb.insert(
                    InlineKeyboardButton('🌆' + product, callback_data=f'change_image_{product}'))
                ikb.insert(
                    InlineKeyboardButton(f'{price}₽', callback_data='change_price_' + product))

        ikb.add(
            InlineKeyboardButton(
                text='◀️',
                callback_data=f'prev_menu_page {page} {next_page_len} {del_product}'
            )
        )

        ikb.insert(InlineKeyboardButton(f'{page}', callback_data='None'))

        ikb.insert(
            InlineKeyboardButton(
                text='▶️',
                callback_data=f'next_menu_page {page} {next_page_len} {del_product}'
            )
        )

        if del_product:
            ikb.add(InlineKeyboardButton('Назад', callback_data='back_to_edit_menu'))
        else:
            ikb.add(InlineKeyboardButton('Удалить товар', callback_data='del_product_page'))
            ikb.insert(InlineKeyboardButton('Добавить товар', callback_data='add_product_page'))
            ikb.add(InlineKeyboardButton('Назад', callback_data='back'))
            ikb.insert(InlineKeyboardButton('Помощь', callback_data='menu_help'))

        return ikb

    @classmethod
    async def show_page(
            cls, user_id: int, msg_id: int, bot: Bot, show_help: bool = False,
            del_product: bool = False, page: int = 1
    ):
        msg_help = MENU_HELP if show_help else ''

        await bot.edit_message_text(
            text=EDIT_MENU_TITLE + msg_help,
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=await cls.get_page(del_product, page)
        )

    @classmethod
    async def change_desc(cls, user_id: int, callback: CallbackQuery, bot: Bot):
        product = callback.data[12:]
        await ChangeProduct.set_product(product)
        await ChangeProduct.get_new_desc.set()

        await bot.send_message(
            chat_id=user_id,
            text='Введите описание товара (состав):',
            reply_markup=ikb_cancel
        )

        await callback.answer('Редактирование описания')

    @classmethod
    async def change_image(cls, user_id: int, callback: CallbackQuery, bot: Bot):
        product = callback.data[13:]
        await ChangeProduct.set_product(product)
        await ChangeProduct.get_new_product_image.set()

        await bot.send_message(
            chat_id=user_id,
            text='Отправьте ссылку на изображение:',
            reply_markup=ikb_cancel
        )

        await callback.answer('Редактирование изображения')

    @classmethod
    async def change_price(cls, user_id: int, callback: CallbackQuery, bot: Bot):
        product = callback.data[13:]
        await ChangeProduct.set_product(product)
        await ChangeProduct.get_new_price.set()

        await bot.send_message(
            chat_id=user_id,
            text='Введите стоимость товара:',
            reply_markup=ikb_cancel
        )

        await callback.answer('Редактирование цены')

    @classmethod
    async def add_product(cls, user_id: int, callback: CallbackQuery, bot: Bot):
        await bot.send_message(
            chat_id=user_id,
            text='Введите название товара:',
            reply_markup=ikb_cancel
        )

        await AddProduct.get_name.set()
        await callback.answer('Добавление товара')

    @classmethod
    async def delete_product(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        product = callback.data[15:]
        await cls.db.delete_product(product)
        await cls.show_page(user_id, msg_id, bot)
        await callback.answer(f'Товар удалён', show_alert=True)

    @classmethod
    async def pagination(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        del_product = True if 'True' in callback.data else False
        data = callback.data.split()
        page = int(data[1])
        next_page_len = int(data[2])

        if 'next' in callback.data and next_page_len > 0:
            await cls.show_page(user_id, msg_id, bot, del_product=del_product, page=page + 1)

        elif 'prev' in callback.data and page != 1:
            await cls.show_page(user_id, msg_id, bot, del_product=del_product, page=page - 1)


class Settings(Menu):

    @staticmethod
    async def get_page() -> InlineKeyboardMarkup:
        ikb = InlineKeyboardMarkup()
        data = await get_json('data.json')
        bot_enabled = data['is_bot_enabled']

        if bot_enabled:
            ikb.add(InlineKeyboardButton('Выключить бота', callback_data='state_bot_off'))
        else:
            ikb.add(InlineKeyboardButton('Включить бота', callback_data='state_bot_on'))

        ikb.add(InlineKeyboardButton('Стартовое изображение', callback_data='change_main_image'))
        ikb.add(InlineKeyboardButton('Назад', callback_data='back'))

        return ikb

    @classmethod
    async def show_page(cls, user_id: int, msg_id: int, bot: Bot):
        await bot.edit_message_text(
            text=SETTINGS_TITLE,
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=await cls.get_page()
        )

    @classmethod
    async def switch_state(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        if 'off' in callback.data:
            await set_json('data.json', {'is_bot_enabled': 0})
            await callback.answer('Приём заказов остановлен', show_alert=True)
        else:
            await set_json('data.json', {'is_bot_enabled': 1})
            await callback.answer('Приём заказов запущен', show_alert=True)

        await bot.edit_message_reply_markup(
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=await cls.get_page()
        )

    @classmethod
    async def change_main_image(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        await bot.edit_message_text(
            text='Отправьте изображение:',
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=ikb_cancel
        )
        await ChangeMainImage.get_main_image.set()
        await callback.answer('Редактирование изображения')


class Mail(Menu):

    @classmethod
    async def get_page(cls) -> InlineKeyboardMarkup:
        mails_count = await cls.db.get_mails_count()
        mails = f'Мои рассылки ({mails_count})' if mails_count else 'Мои рассылки'
        ikb = InlineKeyboardMarkup()
        ikb.add(InlineKeyboardButton('Создать рассылку', callback_data='create_mail'))
        ikb.add(InlineKeyboardButton(mails, callback_data='my_mails'))
        ikb.add(InlineKeyboardButton('Назад', callback_data='back'))
        ikb.insert(InlineKeyboardButton('Помощь', callback_data='mails_help'))

        return ikb

    @classmethod
    async def show_page(cls, user_id: int, msg_id: int, bot: Bot, show_help: bool = False):
        help_msg = MAILS_HELP if show_help else ''

        await bot.edit_message_text(
            text=MAILS_TITLE + help_msg,
            chat_id=user_id,
            message_id=msg_id,
            reply_markup=await cls.get_page()
        )

    @classmethod
    async def my_mails(cls, page: int) -> InlineKeyboardMarkup:
        pages = cls.db.get_mails_count()

        ikb = InlineKeyboardMarkup()
        ikb.add(InlineKeyboardButton('◀️', callback_data=f'prev_my_mails {page} {pages}'))
        ikb.insert(InlineKeyboardButton(f'{page}', callback_data='None'))
        ikb.insert(InlineKeyboardButton('▶️', callback_data=f'next_my_mails {page} {pages}'))
        ikb.add(InlineKeyboardButton('Удалить', callback_data='delete_mail'))
        ikb.add(InlineKeyboardButton('Назад', callback_data='admin_mails'))
        ikb.insert(InlineKeyboardButton('Отправить', callback_data='send_mail'))

        return ikb

    @classmethod
    async def get_mails(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        data = callback.data.split()

        if 'prev' in callback.data:
            page = int(data[1])
            if page == 1:
                return await callback.answer()
            await cls.db.move_selected_mail(False)
            page -= 1
        elif 'next' in callback.data:
            page = int(data[1])
            total_pages = int(data[2])
            if page == total_pages:
                return await callback.answer()
            await cls.db.move_selected_mail(True)
            page += 1
        try:
            mail_id, mail_text = await cls.db.get_mail()

            await bot.edit_message_text(
                text=mail_text,
                chat_id=user_id,
                message_id=msg_id,
                reply_markup=await cls.my_mails(mail_id)
            )
        except TypeError:
            await callback.answer('Список рассылок пуст')

    @classmethod
    async def create(cls, user_id: int, bot: Bot):
        await StateMail.new_mail.set()
        await bot.send_message(
            chat_id=user_id,
            text='Введите текст рассылки',
            reply_markup=ikb_cancel
        )

    @classmethod
    async def delete(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        if 'yes' in callback.data:
            state = await cls.db.delete_mail()
            await callback.answer('Рассылка удалена', show_alert=True)
            if state:
                mail_id, mail_text = await cls.db.get_mail()

                await bot.edit_message_text(
                    text=mail_text,
                    chat_id=user_id,
                    message_id=msg_id,
                    reply_markup=await cls.my_mails(mail_id)
                )
            else:
                await cls.show_page(user_id, msg_id, bot)

        elif 'no' in callback.data:
            await callback.answer('Удаление отменено', show_alert=True)
            await bot.delete_message(user_id, msg_id)
        else:
            await bot.send_message(
                chat_id=user_id,
                text='Удалить рассылку?',
                reply_markup=ikb_del_mail
            )

    @classmethod
    async def send(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        if 'yes' in callback.data:
            users = await cls.db.get_all_user_id()
            _, mail_text = await cls.db.get_mail()
            if users:
                for user in users:
                    await bot.send_message(
                        chat_id=user,
                        text=mail_text
                    )

                word = 'клиенту' if str(len(users))[-1] == '1' else 'клиентам'

                await callback.answer(
                    f'Рассылка успешно отправлена {len(users)} {word}',
                    show_alert=True
                )

                await bot.delete_message(user_id, msg_id)
            else:
                await callback.answer('Список клиентов пуст', show_alert=True)
                await bot.delete_message(user_id, msg_id)

        elif 'no' in callback.data:
            await callback.answer('Отправление отменено', show_alert=True)
            await bot.delete_message(user_id, msg_id)
        else:
            mail_text = await cls.db.get_mail()

            await bot.send_message(
                chat_id=user_id,
                text=SEND_MAIL + mail_text[1],
                reply_markup=ikb_send_mail
            )


class MyOrders(Menu):

    @staticmethod
    async def get_page(
            count: int, selected_page: int, user_orders_count: int, user_orders: list
    ) -> str:

        num = 1
        answer = ''
        rows = 5

        if selected_page == 0:
            or_num = [_ for _ in range(1, user_orders_count + 1)][count * -1:]
            for order_number, price, order_list, date, time in user_orders[count * -1:]:
                answer += f'[{or_num[num - 1]}]  <b>Заказ №<u>{order_number}</u></b>\n' \
                          f'{order_list} | <b>Оплата: {price}₽</b>\n[{date} {time}]\n\n'
                num += 1
        else:
            end = selected_page * rows
            start = end - rows
            or_num = [_ for _ in range(1, user_orders_count + 1)][start:end]
            for order_number, price, order_list, date, time in user_orders[start:end]:
                answer += f'[{or_num[num - 1]}]  <b>Заказ №<u>{order_number}</u></b>\n' \
                          f'{order_list} | <b>Оплата: {price}₽</b>\n[{date} {time}]\n\n'
                num += 1

        await sleep(0.01)

        return answer

    @staticmethod
    async def get_markup(page: int, pages: int) -> InlineKeyboardMarkup:
        ikb = InlineKeyboardMarkup()
        ikb.add(InlineKeyboardButton('◀️', callback_data=f'prev_my_orders {page} {pages}'))
        ikb.insert(InlineKeyboardButton(f'{page}', callback_data='None'))
        ikb.insert(InlineKeyboardButton('▶️', callback_data=f'next_my_orders {page} {pages}'))

        return ikb

    @classmethod
    async def show_page(cls, message, bot: Bot, user_id: int, msg_id: int, selected_page: int = 0):
        user_orders = await cls.db.get_user_orders(message.from_user.id)
        user_orders_count = len(user_orders)

        if user_orders_count != 0:
            if user_orders_count > 5:
                last_page = user_orders_count % 5
                order_pages = user_orders_count // 5
                page = 0 if last_page == 0 else 1
                total_pages = int(order_pages) + page
                if last_page == 0:
                    last_page = 5

                if selected_page == 0:
                    await message.answer(
                        text=await cls.get_page(
                            last_page, selected_page, user_orders_count, user_orders
                        ),
                        reply_markup=await cls.get_markup(total_pages, total_pages)
                    )
                else:
                    await bot.edit_message_text(
                        text=await cls.get_page(
                            last_page, selected_page, user_orders_count, user_orders
                        ),
                        chat_id=user_id,
                        message_id=msg_id,
                        reply_markup=await cls.get_markup(selected_page, total_pages)
                    )
            else:
                if selected_page == 0:
                    await message.answer(cls.get_page(
                        user_orders_count, selected_page, user_orders_count, user_orders
                    )
                                         )
                else:
                    await bot.send_message(
                        chat_id=user_id,
                        text=cls.get_page(
                            user_orders_count, selected_page, user_orders_count, user_orders
                        )
                    )
        else:
            await bot.send_message(
                chat_id=user_id,
                text='Список заказов пуст'
            )

    @classmethod
    async def pagination(cls, user_id: int, msg_id: int, callback: CallbackQuery, bot: Bot):
        data = callback.data.split()
        page = int(data[1])
        if 'prev' in callback.data:
            if page != 1:
                await cls.show_page(callback, bot, user_id, msg_id, page - 1)
        else:
            total_pages = int(data[2])
            if page != total_pages:
                await cls.show_page(callback, bot, user_id, msg_id, page + 1)
