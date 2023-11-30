import json
import random
import string
import hashlib
import datetime

from aiogram import types, Bot
from asyncio import sleep
from openpyxl import Workbook
from openpyxl.styles import Alignment

from order_db import OrderDB
from config import TIME_ZONE, logger
from messages import ERROR_F


@logger.catch
async def set_json(path: str, data: dict):
    json_data = get_json(path)
    json_data = json_data | data
    with open(path, 'w') as file:
        json.dump(json_data, file, ensure_ascii=False)
    await sleep(0.1)


@logger.catch
def get_json(path: str) -> dict:
    with open(path, 'r') as file:
        return json.load(file)


@logger.catch
def gen_password() -> str:
    password = ''
    pw_str = string.digits+string.ascii_uppercase
    for ch in ['I', 'O', '0', 'J', 'Z', 'C']:
        pw_str = pw_str.replace(ch, '')
    length = len(pw_str)
    for i in range(5):
        password += pw_str[random.randint(0, length-1)]
    return password


@logger.catch
async def update_password() -> str:
    pw = gen_password()
    pw_dict = {'employee_password': pw}
    await set_json('data.json', pw_dict)
    return pw


@logger.catch
async def generate_id() -> str:
    hash_data = ''
    for i in range(10):
        hash_data += string.ascii_letters[random.randint(0, len(string.ascii_letters)-1)]
    await sleep(0.1)

    return hashlib.md5(hash_data.encode()).hexdigest()


@logger.catch
async def generate_order_number() -> int:
    order_number = random.randint(100000, 999999)
    while order_number in await OrderDB.get_order_numbers():
        order_number += 1
        if order_number > 999999:
            order_number = random.randint(100000, 999999)

    return order_number


@logger.catch
async def get_time():
    now = datetime.datetime.now() + datetime.timedelta(hours=TIME_ZONE)
    time = now.strftime('%H:%M')
    hour = now.hour
    minute = now.minute
    await sleep(0.1)

    return time, hour, minute


@logger.catch
async def get_24h_orders_list(message):
    orders_24 = await OrderDB.get_orders(1)
    if len(orders_24) == 0:
        await message.answer('Список заказов пуст')
    else:
        answer = ''
        for order_number, price, order_list, comment, date, time in orders_24:
            if comment is not None:
                comment = f'Комментарий: {comment} | '
            else:
                comment = ''
            answer += f'<b>Заказ №<u>{order_number}</u></b> | {order_list} | ' \
                      f'Оплата: {int(price)}₽ | {comment}{date} {time}\n\n'
        await message.answer(answer)


@logger.catch
async def send_order_to_employees(comment: str, order_list: str, bot: Bot, order_number: int, user_time_str: str,
                                  price: int, date: str, time: str):
    if comment is None:
        comm = ''
    else:
        comm = f'\n\n✏ Комментарий: {comment}'

    order = json.loads(order_list.replace('\'', '"'))
    order_str = ''
    for employee in await OrderDB.get_id_by_status('Повар'):
        for product in order:
            order_str += f'\n ▫️ {product}: {order[product]}'
        await bot.send_message(employee,
                               f'<b>Заказ №<u>{order_number}</u></b>'+user_time_str+order_str+f'\n'
                               f'__________\n'
                               f'{price}₽'+comm+f'\n\n'
                               f'{date} {time}')


@logger.catch
async def error_to_db(message: types.Message, bot):
    now = datetime.datetime.now() + datetime.timedelta(hours=TIME_ZONE)
    date = now.strftime('%d.%m.%Y')
    time = await get_time()
    await OrderDB.insert_error(message.from_user.full_name, message.text, date, time[0])
    await bot.send_message(message.from_user.id, ERROR_F)
    await bot.send_message(5765637028, message.from_user.full_name + '\n' + message.text)


@logger.catch
async def get_xlsx() -> str:
    wb = Workbook()
    ws = wb.active
    ws.append(['Номер заказа', 'Заказ', 'Стоимость', 'Дата', 'Время'])

    table = ['A', 'B', 'C', 'D', 'E']

    for ch in table:
        cell = ws[f'{ch}1']
        cell.style = 'Accent1'
        cell.alignment = Alignment(horizontal='center')

    for i, order in enumerate(await OrderDB.get_all_from_archive()):
        ws.append([order[i] for i in [1, 4, 3, 6, 7]])
        ws[f'C{i + 2}'].number_format = '#,## ₽'
        ws[f'D{i + 2}'].alignment = Alignment(horizontal='right')
        ws[f'E{i + 2}'].alignment = Alignment(horizontal='right')

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 10

    now = datetime.datetime.now() + datetime.timedelta(hours=TIME_ZONE)
    now = now.strftime('%d.%m.%Y') + '.xlsx'
    wb.save(now)

    return now


@logger.catch
async def set_answer(count: int, selected_page: int, user_orders_count: int, user_orders: list) -> str:
    num = 1
    answer = ''
    rows = 5

    if selected_page == 0:
        or_num = [_ for _ in range(1, user_orders_count + 1)][count * -1:]
        for order_number, price, order_list, date, time in user_orders[count * -1:]:
            answer += f'[{or_num[num - 1]}]  <b>Заказ №<u>{order_number}</u></b>\n' \
                      f'{order_list} | <b>Оплата: {price}₽</b>\n[{date} {time}]\n\n'
            num += 1
    else:
        end = selected_page * rows
        start = end - rows
        or_num = [_ for _ in range(1, user_orders_count + 1)][start:end]
        for order_number, price, order_list, date, time in user_orders[start:end]:
            answer += f'[{or_num[num - 1]}]  <b>Заказ №<u>{order_number}</u></b>\n' \
                      f'{order_list} | <b>Оплата: {price}₽</b>\n[{date} {time}]\n\n'
            num += 1

    await sleep(0.1)
    return answer
