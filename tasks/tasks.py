import datetime
import os
import asyncio

from openpyxl import Workbook
from openpyxl.styles import Alignment
from celery import Celery
from config import REDIS_HOST, REDIS_PORT, TIME_ZONE
from order_db import OrderDB
from config import bot


celery_event_loop = asyncio.new_event_loop()

celery = Celery('task', broker=f'redis://{REDIS_HOST}:{REDIS_PORT}')
celery.autodiscover_tasks()


async def xlsx(user_id: int):
    await asyncio.sleep(10)
    wb = Workbook()
    ws = wb.active
    ws.append(['Номер заказа', 'Заказ', 'Стоимость', 'Дата', 'Время'])

    table = ['A', 'B', 'C', 'D', 'E']

    for ch in table:
        cell = ws[f'{ch}1']
        cell.style = 'Accent1'
        cell.alignment = Alignment(horizontal='center')

    orders = await OrderDB.get_all_from_orders()

    for i, order in enumerate(orders):
        ws.append([order[i] for i in [1, 4, 3, 6, 7]])
        ws[f'C{i + 2}'].number_format = '#,## ₽'
        ws[f'D{i + 2}'].alignment = Alignment(horizontal='right')
        ws[f'E{i + 2}'].alignment = Alignment(horizontal='right')

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 10

    now = datetime.datetime.now() + datetime.timedelta(hours=TIME_ZONE)
    file = now.strftime('%d.%m.%Y') + '.xlsx'
    wb.save(file)

    await bot.send_document(user_id, open(file, 'rb'))
    for file in os.listdir():
        if '.xlsx' in file:
            os.remove(file)


async def mail(user_id: int):
    await asyncio.sleep(3)
    users = await OrderDB.get_all_user_id()
    _, mail_text = await OrderDB.get_mail()
    if users:
        for user in users:
            await bot.send_message(
                chat_id=user,
                text=mail_text
            )

        word = 'клиенту' if str(len(users))[-1] == '1' else 'клиентам'

        await bot.send_message(user_id, f'Рассылка успешно отправлена {len(users)} {word}')
    else:
        await bot.send_message('Список клиентов пуст')


@celery.task
def get_xlsx(user_id: int):
    celery_event_loop.run_until_complete(xlsx(user_id))


@celery.task
def send_mail(user_id: int):
    celery_event_loop.run_until_complete(mail(user_id))

